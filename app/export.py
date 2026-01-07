# ==============================================
# DIŞA AKTARMA (EXPORT) MODÜLÜ
# ==============================================
# Bu dosya sınav programını PDF ve Excel
# formatlarında dışa aktarma işlemlerini yapar.
# Takvim Formatı: Tarihler sütun, saatler satır
# Her bölüm için ayrı tablo
# PDF: Yer imleri ve içindekiler ile kolay navigasyon
# ==============================================

import os
from datetime import datetime
from collections import defaultdict

# PDF oluşturma için
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    KeepTogether, BaseDocTemplate, PageTemplate, Frame, NextPageTemplate
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus.tableofcontents import TableOfContents
from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel oluşturma için
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.models.exam import get_all_exams
from app.database import execute_query


def get_export_folder():
    """Export klasörünün yolunu döndürür."""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    export_folder = os.path.join(base_dir, 'exports')
    if not os.path.exists(export_folder):
        os.makedirs(export_folder)
    return export_folder


def get_departments():
    """Tüm bölümleri getirir."""
    query = "SELECT id, name FROM departments ORDER BY name"
    return execute_query(query)


def get_exams_by_department(department_id):
    """Bölüme göre sınavları getirir."""
    query = """
        SELECT e.*, 
               c.code as course_code, 
               c.name as course_name,
               c.student_count,
               cl.name as classroom_name,
               s.name as supervisor_name,
               i.name as instructor_name
        FROM exam_schedule e
        LEFT JOIN courses c ON e.course_id = c.id
        LEFT JOIN classrooms cl ON e.classroom_id = cl.id
        LEFT JOIN instructors s ON e.supervisor_id = s.id
        LEFT JOIN instructors i ON c.instructor_id = i.id
        WHERE c.department_id = ?
        ORDER BY e.exam_date, e.start_time
    """
    return execute_query(query, (department_id,))


def organize_exams_as_calendar(exams):
    """
    Sınavları takvim formatında organize eder.
    
    Döndürür:
        dates: Benzersiz tarihler listesi (DD.MM.YYYY formatında)
        time_slots: Benzersiz saat dilimleri listesi
        calendar_data: {(tarih, saat): exam_info} sözlüğü
    """
    # Tarihleri DD.MM.YYYY formatına çevir
    def format_date(date_str):
        if '-' in date_str and len(date_str) == 10:
            parts = date_str.split('-')
            return f"{parts[2]}.{parts[1]}.{parts[0]}"
        return date_str
    
    raw_dates = sorted(set(exam['exam_date'] for exam in exams))
    dates = [format_date(d) for d in raw_dates]
    time_slots = sorted(set(exam['start_time'] + '-' + exam['end_time'] for exam in exams))
    
    calendar_data = {}
    for exam in exams:
        date = format_date(exam['exam_date'])  # Formatlanmış tarih kullan
        time_slot = exam['start_time'] + '-' + exam['end_time']
        key = (date, time_slot)
        
        exam_info = f"{exam['course_code']}\n{exam['classroom_name']}"
        
        # supervisor_name veya instructor_name'i kontrol et
        supervisor = None
        try:
            supervisor = exam['supervisor_name']
        except (KeyError, IndexError):
            pass
        
        if not supervisor:
            try:
                supervisor = exam['instructor_name']
            except (KeyError, IndexError):
                pass
        
        if supervisor:
            exam_info += f"\n({supervisor})"
        
        if key in calendar_data:
            calendar_data[key] += '\n---\n' + exam_info
        else:
            calendar_data[key] = exam_info
    
    return dates, time_slots, calendar_data


def register_fonts():
    """Türkçe karakter desteği için font kaydet."""
    font_name = 'Helvetica'
    font_name_bold = 'Helvetica-Bold'
    font_registered = False
    
    # Windows Arial
    try:
        arial_path = 'C:/Windows/Fonts/arial.ttf'
        arial_bold_path = 'C:/Windows/Fonts/arialbd.ttf'
        
        if os.path.exists(arial_path):
            pdfmetrics.registerFont(TTFont('Arial', arial_path))
            if os.path.exists(arial_bold_path):
                pdfmetrics.registerFont(TTFont('Arial-Bold', arial_bold_path))
                font_name_bold = 'Arial-Bold'
            font_name = 'Arial'
            font_registered = True
    except:
        pass
    
    return font_name, font_name_bold, font_registered


def convert_turkish(text, font_registered):
    """Türkçe karakterleri ASCII'ye çevir (font yoksa)."""
    if font_registered or not text:
        return text
    
    tr_chars = {'ı': 'i', 'İ': 'I', 'ğ': 'g', 'Ğ': 'G', 
               'ü': 'u', 'Ü': 'U', 'ş': 's', 'Ş': 'S',
               'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C'}
    for tr, en in tr_chars.items():
        text = text.replace(tr, en)
    return text


class BookmarkedDocTemplate(BaseDocTemplate):
    """
    Yer imleri (bookmarks) destekleyen PDF şablonu.
    Excel'deki sheet navigasyonu gibi bölümler arası gezinme sağlar.
    """
    def __init__(self, filename, **kwargs):
        self.bookmarks = []
        BaseDocTemplate.__init__(self, filename, **kwargs)
    
    def afterFlowable(self, flowable):
        """Her flowable sonrası yer imi ekle."""
        if hasattr(flowable, '_bookmarkName'):
            # PDF'e yer imi ekle
            self.canv.bookmarkPage(flowable._bookmarkName)
            self.canv.addOutlineEntry(
                flowable._bookmarkTitle,
                flowable._bookmarkName,
                level=flowable._bookmarkLevel
            )


class BookmarkedParagraph(Paragraph):
    """Yer imi destekleyen Paragraph sınıfı."""
    def __init__(self, text, style, bookmarkName=None, bookmarkTitle=None, level=0):
        Paragraph.__init__(self, text, style)
        if bookmarkName:
            self._bookmarkName = bookmarkName
            self._bookmarkTitle = bookmarkTitle or text
            self._bookmarkLevel = level


def export_to_pdf():
    """
    Sınav programını takvim formatında PDF olarak dışa aktarır.
    Her bölüm için ayrı sayfa ve yer imi ile kolay navigasyon.
    Excel'deki sheet yapısı gibi sol panelden bölümler arası geçiş yapılabilir.
    """
    # Dosya oluştur
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = 'sinav_programi_' + timestamp + '.pdf'
    file_path = os.path.join(get_export_folder(), filename)
    
    # Font kaydet
    font_name, font_name_bold, font_registered = register_fonts()
    
    # PDF belgesi oluştur (Yer imli şablon)
    doc = BookmarkedDocTemplate(
        file_path,
        pagesize=landscape(A4),
        rightMargin=0.5*cm,
        leftMargin=0.5*cm,
        topMargin=1.2*cm,
        bottomMargin=1*cm
    )
    
    # Sayfa şablonu
    frame = Frame(
        doc.leftMargin, doc.bottomMargin,
        doc.width, doc.height,
        id='normal'
    )
    
    def add_page_header_footer(canvas, doc):
        """Her sayfaya başlık ve sayfa numarası ekle."""
        canvas.saveState()
        
        # Üst başlık çizgisi
        canvas.setStrokeColor(colors.HexColor('#2563eb'))
        canvas.setLineWidth(2)
        canvas.line(0.5*cm, landscape(A4)[1] - 0.8*cm, landscape(A4)[0] - 0.5*cm, landscape(A4)[1] - 0.8*cm)
        
        # Üst başlık metni
        canvas.setFont(font_name_bold, 10)
        canvas.setFillColor(colors.HexColor('#2563eb'))
        canvas.drawString(0.7*cm, landscape(A4)[1] - 0.6*cm, "SINAV PROGRAMI")
        
        # Tarih (sağ üst)
        today = datetime.now().strftime('%d.%m.%Y')
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor('#666666'))
        canvas.drawRightString(landscape(A4)[0] - 0.7*cm, landscape(A4)[1] - 0.6*cm, f"Tarih: {today}")
        
        # Alt sayfa numarası
        canvas.setFont(font_name, 9)
        canvas.setFillColor(colors.HexColor('#666666'))
        page_num = canvas.getPageNumber()
        canvas.drawCentredString(landscape(A4)[0] / 2, 0.5*cm, f"Sayfa {page_num}")
        
        # Alt çizgi
        canvas.setStrokeColor(colors.HexColor('#e5e7eb'))
        canvas.setLineWidth(1)
        canvas.line(0.5*cm, 0.8*cm, landscape(A4)[0] - 0.5*cm, 0.8*cm)
        
        canvas.restoreState()
    
    template = PageTemplate(id='main', frames=frame, onPage=add_page_header_footer)
    doc.addPageTemplates([template])
    
    styles = getSampleStyleSheet()
    
    # Kapak başlık stili
    cover_title_style = ParagraphStyle(
        'CoverTitle',
        parent=styles['Heading1'],
        fontName=font_name_bold,
        fontSize=28,
        alignment=TA_CENTER,
        spaceAfter=20,
        textColor=colors.HexColor('#1e40af')
    )
    
    cover_subtitle_style = ParagraphStyle(
        'CoverSubtitle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#666666'),
        spaceAfter=40
    )
    
    # İçindekiler stili
    toc_title_style = ParagraphStyle(
        'TOCTitle',
        parent=styles['Heading1'],
        fontName=font_name_bold,
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20,
        textColor=colors.HexColor('#1e40af')
    )
    
    toc_item_style = ParagraphStyle(
        'TOCItem',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=11,
        leftIndent=20,
        spaceBefore=8,
        spaceAfter=8,
        textColor=colors.HexColor('#2563eb')
    )
    
    # Bölüm başlık stili
    dept_style = ParagraphStyle(
        'DeptTitle',
        parent=styles['Heading2'],
        fontName=font_name_bold,
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=15,
        spaceBefore=5,
        borderColor=colors.HexColor('#2563eb'),
        borderWidth=1,
        borderPadding=5
    )
    
    elements = []
    
    # Bölümleri al
    departments = get_departments()
    dept_with_exams = []
    
    # Önce sınavı olan bölümleri belirle
    for dept in departments:
        exams = get_exams_by_department(dept['id'])
        if len(exams) > 0:
            dept_with_exams.append({
                'id': dept['id'],
                'name': dept['name'],
                'exam_count': len(exams)
            })
    
    if not dept_with_exams:
        # Sınav yoksa basit PDF
        elements.append(Paragraph("Henuz planlanmis sinav bulunmamaktadir.", styles['Normal']))
        doc.build(elements)
        return file_path
    
    # =====================
    # KAPAK SAYFASI
    # =====================
    elements.append(Spacer(1, 4*cm))
    
    cover_title = BookmarkedParagraph(
        "SINAV PROGRAMI",
        cover_title_style,
        bookmarkName="cover",
        bookmarkTitle="Kapak",
        level=0
    )
    elements.append(cover_title)
    
    today = datetime.now().strftime('%d.%m.%Y')
    semester_info = "2024-2025 Akademik Yili"
    elements.append(Paragraph(semester_info, cover_subtitle_style))
    elements.append(Paragraph(f"Olusturulma Tarihi: {today}", cover_subtitle_style))
    
    elements.append(Spacer(1, 2*cm))
    
    # Özet istatistikler
    total_exams = sum(d['exam_count'] for d in dept_with_exams)
    stats_data = [
        ['ISTATISTIKLER', ''],
        ['Toplam Bolum Sayisi', str(len(dept_with_exams))],
        ['Toplam Sinav Sayisi', str(total_exams)],
    ]
    
    stats_table = Table(stats_data, colWidths=[8*cm, 4*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('SPAN', (0, 0), (1, 0)),
    ]))
    elements.append(stats_table)
    
    elements.append(PageBreak())
    
    # =====================
    # İÇİNDEKİLER SAYFASI
    # =====================
    toc_title = BookmarkedParagraph(
        "ICINDEKILER",
        toc_title_style,
        bookmarkName="toc",
        bookmarkTitle="Icindekiler",
        level=0
    )
    elements.append(toc_title)
    elements.append(Spacer(1, 0.5*cm))
    
    # İçindekiler tablosu
    toc_data = [['No', 'Bolum Adi', 'Sinav Sayisi']]
    for i, dept in enumerate(dept_with_exams, 1):
        dept_name_clean = convert_turkish(dept['name'], font_registered)
        toc_data.append([str(i), dept_name_clean, str(dept['exam_count'])])
    
    toc_table = Table(toc_data, colWidths=[1.5*cm, 16*cm, 3*cm])
    toc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    elements.append(toc_table)
    
    elements.append(Spacer(1, 1*cm))
    
    # Navigasyon ipucu
    nav_hint = Paragraph(
        "<i>Not: PDF goruntuleyicinizde sol panelden (Yer Imleri/Bookmarks) bolumler arasi hizli gecis yapabilirsiniz.</i>",
        ParagraphStyle('NavHint', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#666666'), alignment=TA_CENTER)
    )
    elements.append(nav_hint)
    
    elements.append(PageBreak())
    
    # =====================
    # BÖLÜM SAYFALARI
    # =====================
    for idx, dept in enumerate(dept_with_exams):
        dept_id = dept['id']
        dept_name = convert_turkish(dept['name'], font_registered)
        
        # Bu bölümün sınavlarını al
        exams = get_exams_by_department(dept_id)
        
        # Bölüm başlığı (yer imli)
        bookmark_name = f"dept_{dept_id}"
        dept_title = BookmarkedParagraph(
            f"{idx + 1}. {dept_name}",
            dept_style,
            bookmarkName=bookmark_name,
            bookmarkTitle=dept_name,
            level=1
        )
        elements.append(dept_title)
        
        # Bölüm bilgisi
        dept_info = Paragraph(
            f"<font color='#666666'>Toplam {len(exams)} sinav</font>",
            ParagraphStyle('DeptInfo', parent=styles['Normal'], fontSize=9, spaceAfter=10)
        )
        elements.append(dept_info)
        
        # Takvim formatına dönüştür
        dates, time_slots, calendar_data = organize_exams_as_calendar(exams)
        
        if not dates or not time_slots:
            continue
        
        # Tablo verisi oluştur
        header_row = ['Saat'] + dates
        table_data = [header_row]
        
        # Saat satırları
        for time_slot in time_slots:
            row = [time_slot]
            for date in dates:
                key = (date, time_slot)
                cell_value = calendar_data.get(key, '')
                cell_value = convert_turkish(cell_value, font_registered)
                row.append(cell_value)
            table_data.append(row)
        
        # Sütun genişlikleri
        num_cols = len(dates) + 1
        available_width = 26*cm
        col_width = available_width / num_cols
        col_widths = [col_width] * num_cols
        col_widths[0] = min(2.5*cm, col_width)  # Saat sütunu
        
        # Tablo oluştur
        table = Table(table_data, colWidths=col_widths)
        
        # Alternatif satır renkleri için stil
        table_style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#e5e7eb')),
            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
            ('FONTNAME', (0, 1), (0, -1), font_name_bold),
            ('FONTNAME', (1, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]
        
        # Sınav olan hücrelere arka plan rengi
        for row_idx, time_slot in enumerate(time_slots, 1):
            for col_idx, date in enumerate(dates, 1):
                key = (date, time_slot)
                if key in calendar_data and calendar_data[key]:
                    table_style_cmds.append(
                        ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#dbeafe'))
                    )
        
        table.setStyle(TableStyle(table_style_cmds))
        elements.append(table)
        
        # Her bölüm yeni sayfada
        if idx < len(dept_with_exams) - 1:
            elements.append(PageBreak())
    
    doc.build(elements)
    return file_path


def export_to_excel():
    """
    Sınav programını takvim formatında Excel olarak dışa aktarır.
    Her bölüm için ayrı sayfa.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = 'sinav_programi_' + timestamp + '.xlsx'
    file_path = os.path.join(get_export_folder(), filename)
    
    workbook = Workbook()
    
    # İlk sayfayı sil
    default_sheet = workbook.active
    
    # Stiller
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    time_fill = PatternFill(start_color='e5e7eb', end_color='e5e7eb', fill_type='solid')
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Bölümleri al
    departments = get_departments()
    sheet_created = False
    
    for dept in departments:
        dept_id = dept['id']
        dept_name = dept['name'][:31]  # Excel sayfa adı limiti
        
        # Bu bölümün sınavlarını al
        exams = get_exams_by_department(dept_id)
        
        if len(exams) == 0:
            continue
        
        # Yeni sayfa oluştur
        if not sheet_created:
            sheet = default_sheet
            sheet.title = dept_name
            sheet_created = True
        else:
            sheet = workbook.create_sheet(title=dept_name)
        
        # Takvim formatına dönüştür
        dates, time_slots, calendar_data = organize_exams_as_calendar(exams)
        
        if not dates or not time_slots:
            continue
        
        # Başlık satırı
        sheet.cell(row=1, column=1, value='Saat')
        sheet.cell(row=1, column=1).font = header_font
        sheet.cell(row=1, column=1).fill = header_fill
        sheet.cell(row=1, column=1).alignment = cell_alignment
        sheet.cell(row=1, column=1).border = thin_border
        
        for col_num, date in enumerate(dates, 2):
            cell = sheet.cell(row=1, column=col_num, value=date)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = cell_alignment
            cell.border = thin_border
        
        # Saat satırları
        for row_num, time_slot in enumerate(time_slots, 2):
            # Saat sütunu
            time_cell = sheet.cell(row=row_num, column=1, value=time_slot)
            time_cell.font = Font(bold=True)
            time_cell.fill = time_fill
            time_cell.alignment = cell_alignment
            time_cell.border = thin_border
            
            # Tarih sütunları
            for col_num, date in enumerate(dates, 2):
                key = (date, time_slot)
                cell_value = calendar_data.get(key, '')
                
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # Sınav varsa arka plan rengi
                if cell_value:
                    cell.fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
        
        # Sütun genişlikleri
        sheet.column_dimensions['A'].width = 12
        for col_num in range(2, len(dates) + 2):
            col_letter = chr(64 + col_num)
            sheet.column_dimensions[col_letter].width = 18
        
        # Satır yükseklikleri
        for row_num in range(1, len(time_slots) + 2):
            sheet.row_dimensions[row_num].height = 50
    
    # Hiç sayfa oluşturulmadıysa
    if not sheet_created:
        default_sheet.title = "Boş"
        default_sheet.cell(row=1, column=1, value="Henüz planlanmış sınav bulunmamaktadır.")
    
    workbook.save(file_path)
    return file_path
